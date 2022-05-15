from django.shortcuts import render, redirect
from .models import Document
from openpyxl import Workbook # 엑셀을 만드는 api (엑셀 미설치 시에도 동작)
from io import BytesIO # 엑셀 파일을 전송 할 수 있도록 바이트 배열로 변환
from django.core.files.storage import FileSystemStorage
import pandas as pd
from datetime import datetime
from django.conf import settings
# Create your views here.

def postlist(request):
    documents = Document.objects.all()

    return render(request, "post/postlist.html", context={
        "files": documents
    })

def uploadFile(request):
    return render(request,"post/upload-file.html")

def saveFile(request):
    if request.method == "POST":
        # Fetching the form data
        fileTitle = request.POST["fileTitle"]
        uploadedFile = request.FILES["uploadedFile"]

        # Saving the information in the database
        document = Document(
            title=fileTitle,
            uploadedFile=uploadedFile
        )
        document.save()
    return redirect('post:postlist')
   
def file(request):
    if request.method == "POST":
        fileTitle = request.POST["fileTitle"]
        uploadedFile = request.FILES["uploadedFile"]
        
        # n_path=settings.MEDIA_ROOT+'/'+datetime.today().strftime("%Y-%m-%d")
        
        fs=FileSystemStorage()
        
        filename=fs.save(uploadedFile.name,uploadedFile)
        url=fs.url(filename)
        
        print(filename)
        print(url)

        exceldata=pd.read_excel("."+url)

        # # Saving the information in the database
        # document = Document(
        #     title=fileTitle,
        #     uploadedFile=uploadedFile
        # )
        # document.save()
    return redirect('post:postlist')

# 내가 짜다가 만 report view - 멘토님이 보내주신 엑셀 파일 기반으로 대충 해봄
# 중간의 Report는 model에 생성할 db에 들어갈것 
""" 
def report(request):
    wb = Workbook()  # 엑셀 생성
    ws = wb.active	# 엑셀 활성화
    excelfile = BytesIO() #바이트 배열 생성

    ws['A1']= 'date' # 엑셀 a1 열 이름 정하기
    ws['B1']= 'PO_name'
    ws['C1']= 'address'
    ws['D1']= 'collector'
    ws['E1']= 'quantity'
    ws['F1']= 'converted_qty'
    ws['G1']= 'conllecting_company'


    for i in text: # text 는 db 에 저장된 내용 전체
        content=[i.company,i.product_name,i.count]   #리스트 형태로 1 행씩 생성(a1, b1, c1) 에 각각
        ws.append(content) # 엑셀에 1행을 추가

    wb.close()  #엑셀 닫기
    wb.save(excelfile) # 바이트배열로 저장 (mail 전송 하려면 바이트형태로 변환 되어야 함)

    # 엑셀을 input tag 로 입력받아 db에 저장 하는 코드
    request.FILES['files'].save_to_database(
        model = Report, 	#매핑 대상 db
        mapdict= ['date', 'PO_name', 'address', 'collector', 'quantity', 'converted_qty', 'conllecting_company'  ]) # 칼럼에 매핑하기
"""

